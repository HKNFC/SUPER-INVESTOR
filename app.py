"""
Stock Screener — Streamlit Application

Main entry point for the stock screening dashboard. Provides two tabs:
  - Hisse Tarama: fetch, score, filter, and display stocks ranked by score
  - Backtest: simulate historical performance of screening strategies

Sidebar controls screening parameters; backtest has its own inline controls.
"""

import os
import streamlit as st
import pandas as pd
import numpy as np
import time
from datetime import date, timedelta
from typing import Optional
from config import (
    SUPPORTED_MARKETS, DEFAULT_TOP_N,
    TWELVE_DATA_API_KEY, BENCHMARK_INDEX,
    CACHE_TTL_MARKET_DATA, REQUIRED_FIELDS_FOR_SCORING,
    BIST100_TICKERS, BIST_SEGMENTS,
    SP500_TICKERS, MIDCAP400_TICKERS, USA_SEGMENTS,
)
from data_model import validate_dataframe
from data_fetcher import (
    fetch_market_data, get_last_diagnostics, refresh_eod_cache,
    fetch_backtest_data,
)
from scoring_engine import compute_rs_scores, get_score_breakdown
from institutional_score import STRATEGY_PROFILES, get_debug_info, BLOCK_LABELS
from filters import (
    apply_preset_filter, rank_and_limit,
    get_preset_names, get_preset_info,
)
from utils import format_number, format_percentage, format_large_number, format_market_cap, format_pct_value, is_na
from watchlist import (
    get_watchlist, get_watchlist_tickers, is_in_watchlist,
    add_to_watchlist, remove_from_watchlist, clear_watchlist,
    update_watchlist_scores, export_watchlist_csv,
)
from scan_history import (
    add_scan_entry, add_backtest_entry,
    get_history, delete_entry, clear_history,
)

# Rebalance scheduler'ı arka planda başlat (tek seferlik)
if "scheduler_started" not in st.session_state:
    try:
        from rebalance_scheduler import start_background_scheduler
        start_background_scheduler()
    except Exception:
        pass
    st.session_state["scheduler_started"] = True

st.set_page_config(
    page_title="Hisse Tarayıcı",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

SCAN_MODES = {
    "standard": "Standart Tarama",
    "smart_money": "Sadece Akıllı Para Girenler",
    "early_accumulation": "Erken Accumulation Yakalama",
}
SCAN_MODE_DESCRIPTIONS = {
    "standard": "Tüm skorlarla standart tarama.",
    "smart_money": "Akıllı para girişi tespit edilen hisseler.",
    "early_accumulation": "Erken birikim sinyali veren hisseler.",
}
QUALITY_LABELS = {"none": "Kapalı", "basic": "Temel", "strict": "Sıkı"}
SORT_OPTIONS = {
    "rs_score": "RS Score",
    "technical_score": "Technical Score",
    "combined_score": "Combined Score",
    "institutional_score": "Institutional Score",
    "selection_score": "Selection Score",
    "timing_score": "Timing Score",
}
REBALANCE_LABELS = {
    "1w": "1 Hafta",
    "15d": "15 Gün",
    "1m": "1 Ay",
}
WEIGHT_LABELS = {
    "equal": "Eşit Ağırlık",
}


@st.cache_data(ttl=CACHE_TTL_MARKET_DATA, show_spinner=False)
def _cached_fetch(market: str, _cache_bust: int, skip_fundamentals: bool = False) -> pd.DataFrame:
    return fetch_market_data(market, skip_fundamentals=skip_fundamentals)


def _fmt_rule(rule_key: str, threshold: float) -> str:
    _pct_keys = {"roic_gt", "revenue_growth_gt", "net_margin_gt", "return_12m_gt"}
    labels = {
        "equity_gt": "Özkaynak > {:.0f}",
        "net_income_gt": "Net Kâr > {:.0f}",
        "roic_gt": "ROIC > {:.0f}%",
        "revenue_growth_gt": "Gelir Büyümesi > {:.0f}%",
        "net_margin_gt": "Net Marj > {:.0f}%",
        "debt_to_equity_lt": "B/Ö < {:.1f}",
        "peg_gt": "PEG > {:.0f}",
        "pe_gt": "F/K > {:.0f}",
        "return_12m_gt": "12A Getiri > {:.0f}%",
        "avg_volume_20d_gte": "Ort Hacim >= {:.0f}",
    }
    if rule_key in labels:
        fmt = labels[rule_key]
        val = threshold * 100 if rule_key in _pct_keys else threshold
        return fmt.format(val)
    return f"{rule_key}: {threshold}"


def _score_fmt(val) -> str:
    if is_na(val):
        return "N/A"
    return f"{val:.1f}"


def _missing_metric_warnings(stock_row: pd.Series) -> list:
    return [f for f in REQUIRED_FIELDS_FOR_SCORING if is_na(stock_row.get(f))]


def _render_detail(stock_row: pd.Series, scored_df: Optional[pd.DataFrame] = None) -> None:
    bd = get_score_breakdown(stock_row)

    missing = _missing_metric_warnings(stock_row)
    if missing:
        readable = [f.replace("_", " ").title() for f in missing]
        st.warning(f"Eksik veri: {', '.join(readable)} — skorlar daha az güvenilir olabilir")

    st.markdown(
        f"**{stock_row.get('company_name', '')}** · "
        f"{stock_row.get('sector', '')} · "
        f"{stock_row.get('industry', '')} · "
        f"Piyasa Değeri: {format_market_cap(stock_row.get('market_cap'))}"
    )

    price_val = stock_row.get("price")
    price_str = f"${price_val:,.2f}" if price_val is not None and np.isfinite(price_val) else "N/A"
    rs_val = bd.get("rs_score")
    cat_val = bd.get("rs_category", "N/A")

    tech_val = stock_row.get("technical_score")
    comb_val = stock_row.get("combined_score")
    setup_val = stock_row.get("setup_label", "N/A")

    h1, h2, h3, h4, h5, h6 = st.columns(6)
    h1.metric("Fiyat", price_str)
    h2.metric("RS Skoru", _score_fmt(rs_val))
    h3.metric("Teknik Skor", _score_fmt(tech_val))
    h4.metric("Kombine Skor", _score_fmt(comb_val))
    inst_val = stock_row.get("institutional_score")
    inst_cat = stock_row.get("inst_category", "N/A")
    h5.metric("Kurumsal Skor", _score_fmt(inst_val))
    h6.metric("Kurulum", setup_val)

    st.markdown("---")

    s1, s2, s3, s4, s5 = st.columns(5)
    s1.metric("Finansal Güç", _score_fmt(bd.get("financial_strength")))
    s2.metric("Büyüme", _score_fmt(bd.get("growth")))
    s3.metric("Marj Kalitesi", _score_fmt(bd.get("margin_quality")))
    s4.metric("Değerleme", _score_fmt(bd.get("valuation")))
    s5.metric("Momentum", _score_fmt(bd.get("momentum")))

    i1, i2, i3, i4, i5, i6, i7, i8 = st.columns(8)
    i1.metric("Kurumsal Kat.", inst_cat)
    i2.metric("Seçim Skoru", _score_fmt(stock_row.get("selection_score")))
    i3.metric("Zamanlama Sk.", _score_fmt(stock_row.get("timing_score")))
    i4.metric("K.Kalite", _score_fmt(stock_row.get("inst_quality")))
    i5.metric("K.Büyüme", _score_fmt(stock_row.get("inst_growth")))
    i6.metric("K.Değerleme", _score_fmt(stock_row.get("inst_valuation")))
    i7.metric("K.Momentum", _score_fmt(stock_row.get("inst_momentum")))
    i8.metric("K.Akış", _score_fmt(stock_row.get("inst_flow")))

    ticker = stock_row.get("ticker", "")
    if scored_df is not None:
        dbg = get_debug_info(scored_df, ticker)
        if dbg:
            with st.expander("Kurumsal Skor Debug", expanded=False):
                profile_name = scored_df.attrs.get("_inst_profile", "standard")
                profile_label = STRATEGY_PROFILES.get(profile_name, {}).get("label", profile_name)
                st.caption(f"Profil: **{profile_label}**")
                for block_name, block_info in dbg.items():
                    label = BLOCK_LABELS.get(block_name, block_name)
                    score_str = _score_fmt(block_info.get("score"))
                    used = block_info.get("used", [])
                    missing_m = block_info.get("missing", [])
                    if not used and not missing_m:
                        st.text(f"{label}: {score_str}")
                    else:
                        status_icon = "🟢" if not missing_m else ("🟡" if used else "🔴")
                        st.markdown(f"{status_icon} **{label}**: {score_str}")
                        if used:
                            st.caption(f"  Kullanılan: {', '.join(used)}")
                        if missing_m:
                            st.caption(f"  Eksik: {', '.join(missing_m)}")

    st.markdown("---")

    raw_col, derived_col = st.columns(2)

    with raw_col:
        st.markdown("**Ham Metrikler**")
        st.text(f"Özkaynak:      {format_large_number(stock_row.get('equity'))}")
        st.text(f"Toplam Borç:   {format_large_number(stock_row.get('total_debt'))}")
        st.text(f"Toplam Varlık: {format_large_number(stock_row.get('total_assets'))}")
        st.text(f"Gelir:         {format_large_number(stock_row.get('revenue'))}")
        st.text(f"Net Kâr:       {format_large_number(stock_row.get('net_income'))}")
        st.text(f"F/K:           {format_number(stock_row.get('pe'))}")
        st.text(f"PD/DD:         {format_number(stock_row.get('pb'))}")
        st.text(f"FD/FAVÖK:      {format_number(stock_row.get('ev_ebitda'))}")
        st.text(f"PEG:           {format_number(stock_row.get('peg'))}")

    with derived_col:
        st.markdown("**Türetilmiş Metrikler**")
        st.text(f"B/Ö Oranı:      {format_number(bd.get('debt_to_equity'))}")
        st.text(f"Özkaynak/Varlık:{format_percentage(bd.get('equity_to_assets'))}")
        st.text(f"ROIC:           {format_percentage(bd.get('roic'))}")
        st.text(f"Brüt Marj:      {format_percentage(bd.get('gross_margin'))}")
        st.text(f"Faaliyet Marjı: {format_percentage(bd.get('operating_margin'))}")
        st.text(f"Net Marj:       {format_percentage(bd.get('net_margin'))}")
        st.text(f"Gelir Büy. YoY: {format_percentage(bd.get('revenue_growth'))}")
        st.text(f"Kâr Büy. YoY:   {format_percentage(bd.get('earnings_growth'))}")
        st.text(f"Gelir YBBO 3Y:  {format_percentage(bd.get('revenue_cagr_3y'))}")
        st.text(f"HBK YBBO 3Y:    {format_percentage(bd.get('eps_cagr_3y'))}")

    st.markdown("---")

    r1, r2, r3 = st.columns(3)
    ret_3m = stock_row.get("return_3m")
    ret_6m = stock_row.get("return_6m")
    ret_12m = stock_row.get("return_12m")
    r1.metric("3A Getiri", format_pct_value(ret_3m))
    r2.metric("6A Getiri", format_pct_value(ret_6m))
    r3.metric("12A Getiri", format_pct_value(ret_12m))

    if "price_data" in stock_row and isinstance(stock_row["price_data"], pd.DataFrame):
        price_df = stock_row["price_data"]
        if not price_df.empty and "datetime" in price_df.columns and "close" in price_df.columns:
            st.markdown("**Fiyat Geçmişi**")
            chart_data = price_df.set_index("datetime")[["close"]].rename(columns={"close": "Kapanış"})
            st.line_chart(chart_data, height=250)


def _render_diagnostics(scan_df: Optional[pd.DataFrame] = None) -> None:
    diag = get_last_diagnostics()
    if diag is None:
        return

    with st.expander("Tanılama", expanded=False):
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Çekilen", f"{diag.fetched_tickers}/{diag.total_tickers}")
        d2.metric("Başarısız", str(diag.failed_tickers))
        d3.metric("Eksik Veri", str(diag.incomplete_rows))
        d4.metric("Son Güncelleme", diag.timestamp_str)

        info_parts = []
        if diag.used_mock:
            info_parts.append("Demo/örnek veri kullanılıyor")
        if diag.fallback_triggered:
            info_parts.append("Yedek tetiklendi — API veri döndürmedi")
        info_parts.append(f"Çekme süresi: {diag.duration_seconds:.1f}s")

        cache_stats = st.session_state.get("last_cache_stats")
        if cache_stats:
            info_parts.append(
                f"EOD Cache: {cache_stats.get('fresh', 0)} güncel, "
                f"{cache_stats.get('updated', 0)} güncellendi, "
                f"{cache_stats.get('failed', 0)} başarısız"
            )
        st.caption(" · ".join(info_parts))

        if diag.failed_symbols:
            st.error(f"Başarısız hisseler: {', '.join(diag.failed_symbols)}")

        if hasattr(diag, "fundamentals_total") and diag.fundamentals_total > 0:
            fund_pct = round(diag.fundamentals_with_data / diag.fundamentals_total * 100) if diag.fundamentals_total else 0
            fund_msg = f"Temel Veri: {diag.fundamentals_with_data}/{diag.fundamentals_total} hisse (%{fund_pct})"
            if hasattr(diag, "provider_distribution") and diag.provider_distribution:
                prov_parts = [f"{k}: {v}" for k, v in diag.provider_distribution.items()]
                fund_msg += f" | Kaynak: {' · '.join(prov_parts)}"
            if fund_pct >= 80:
                st.success(fund_msg)
            elif fund_pct >= 50:
                st.info(fund_msg)
            else:
                st.warning(fund_msg)

        if diag.missing_fields_summary:
            missing_lines = [f"{k.replace('_', ' ').title()}: {v} hisse" for k, v in sorted(diag.missing_fields_summary.items(), key=lambda x: -x[1])]
            st.warning("Evrende eksik alanlar: " + " · ".join(missing_lines))

        if diag.errors:
            with st.expander("Hata detayları", expanded=False):
                for err in diag.errors[:20]:
                    st.text(err)

        if scan_df is not None and not scan_df.empty:
            debug_cols = ["ticker", "data_provider"]
            base_fields = ["revenue", "net_income", "equity", "total_debt",
                           "total_assets", "pe", "pb", "revenue_growth",
                           "net_margin", "debt_to_equity", "roic_approx"]
            flag_fields = ["quality_profitable", "quality_growing", "quality_solvent"]
            for c in base_fields + flag_fields:
                if c in scan_df.columns:
                    debug_cols.append(c)
            available = [c for c in debug_cols if c in scan_df.columns]
            if len(available) > 1:
                with st.expander("Veri Sağlayıcı Detayları", expanded=False):
                    if "data_provider" in scan_df.columns:
                        provider_counts = scan_df["data_provider"].value_counts()
                        parts = [f"{k}: {v}" for k, v in provider_counts.items()]
                        st.caption("Sağlayıcı dağılımı: " + " · ".join(parts))
                    st.dataframe(scan_df[available].head(50), use_container_width=True, hide_index=True)


with st.sidebar:
    st.header("Hisse Tarayıcı")

    # ── Rebalans Tarihi Hesaplayıcı ─────────────────────────────────────
    from rebalance_utils import next_rebalance_date, trading_days_until, holiday_name
    from datetime import date as _date

    st.markdown("---")
    st.markdown("**Rebalans Takibi**")
    import json as _json, os as _os
    _reb_save_path = _os.path.join(_os.path.dirname(__file__), ".last_rebalance_date.json")
    _freq_options = {"1m": "1 Ay (21 iş günü)", "15d": "15 Gün (11 iş günü)", "1w": "1 Hafta (5 iş günü)"}

    def _load_reb_date(key, default_days=21):
        d = _date.today() - __import__('datetime').timedelta(days=default_days)
        if _os.path.exists(_reb_save_path):
            try:
                _saved = _json.load(open(_reb_save_path))
                d = _date.fromisoformat(_saved.get(key, str(d)))
            except Exception:
                pass
        return d

    def _save_reb_date(key, val):
        data = {}
        if _os.path.exists(_reb_save_path):
            try:
                data = _json.load(open(_reb_save_path))
            except Exception:
                pass
        data[key] = str(val)
        try:
            with open(_reb_save_path, "w") as _wf:
                _json.dump(data, _wf)
        except Exception:
            pass

    # ── BIST Rebalans ──────────────────────────────────────────────────
    st.markdown("🇹🇷 **BIST İçin Rebalance**")
    _bist_freq_key = st.selectbox(
        "BIST Periyot",
        options=list(_freq_options.keys()),
        format_func=lambda x: _freq_options[x],
        key="reb_freq_bist",
    )
    _bist_last_reb = st.date_input(
        "BIST İçin Rebalance Tarihi",
        value=_load_reb_date("bist_last_reb"),
        key="last_reb_date_bist",
    )
    _save_reb_date("bist_last_reb", _bist_last_reb)
    _bist_next = next_rebalance_date(_bist_last_reb, freq=_bist_freq_key)
    _today = _date.today()
    _bist_days = trading_days_until(_bist_next, from_date=_today)
    if _today >= _bist_next:
        st.error(f"BIST Rebalans: **{_bist_next.strftime('%d.%m.%Y')}** — BUGÜN veya GEÇTİ!")
    elif _bist_days <= 3:
        st.warning(f"BIST Rebalans: **{_bist_next.strftime('%d.%m.%Y')}** — {_bist_days} iş günü kaldı")
    else:
        st.success(f"BIST Rebalans: **{_bist_next.strftime('%d.%m.%Y')}** — {_bist_days} iş günü kaldı")
    _hn_bist = holiday_name(_bist_next)
    if _hn_bist:
        _fd = _bist_next
        while holiday_name(_fd):
            _fd += __import__('datetime').timedelta(days=1)
        st.caption(f"{_bist_next.strftime('%d.%m.%Y')} tatil ({_hn_bist}) → İlk iş günü: **{_fd.strftime('%d.%m.%Y')}**")

    st.markdown("---")

    # ── USA Rebalans ───────────────────────────────────────────────────
    st.markdown("🇺🇸 **USA İçin Rebalance**")
    _usa_freq_key = st.selectbox(
        "USA Periyot",
        options=list(_freq_options.keys()),
        format_func=lambda x: _freq_options[x],
        key="reb_freq_usa",
    )
    _usa_last_reb = st.date_input(
        "USA İçin Rebalance Tarihi",
        value=_load_reb_date("usa_last_reb"),
        key="last_reb_date_usa",
    )
    _save_reb_date("usa_last_reb", _usa_last_reb)
    _usa_next = next_rebalance_date(_usa_last_reb, freq=_usa_freq_key)
    _usa_days = trading_days_until(_usa_next, from_date=_today)
    if _today >= _usa_next:
        st.error(f"USA Rebalans: **{_usa_next.strftime('%d.%m.%Y')}** — BUGÜN veya GEÇTİ!")
    elif _usa_days <= 3:
        st.warning(f"USA Rebalans: **{_usa_next.strftime('%d.%m.%Y')}** — {_usa_days} iş günü kaldı")
    else:
        st.success(f"USA Rebalans: **{_usa_next.strftime('%d.%m.%Y')}** — {_usa_days} iş günü kaldı")
    _hn_usa = holiday_name(_usa_next)
    if _hn_usa:
        _fd = _usa_next
        while holiday_name(_fd):
            _fd += __import__('datetime').timedelta(days=1)
        st.caption(f"{_usa_next.strftime('%d.%m.%Y')} tatil ({_hn_usa}) → İlk iş günü: **{_fd.strftime('%d.%m.%Y')}**")

    st.markdown("---")

    # ── Telegram Bildirim Ayarları ────────────────────────────────────
    st.markdown("📲 **Telegram Bildirimleri**")
    try:
        from telegram_notifier import load_config as _tg_load, save_config as _tg_save, send_message as _tg_send, build_rebalance_message as _tg_build

        _tg_cfg = _tg_load()

        _tg_enabled = st.checkbox(
            "Rebalance Bildirimlerini Etkinleştir",
            value=bool(_tg_cfg.get("enabled", False)),
            key="tg_enabled",
        )
        _tg_token = st.text_input(
            "Bot Token",
            value=_tg_cfg.get("bot_token", ""),
            type="password",
            key="tg_token",
            help="@BotFather'dan alınan token",
        )
        _tg_chat = st.text_input(
            "Chat ID",
            value=_tg_cfg.get("chat_id", ""),
            key="tg_chat",
            help="@userinfobot ile öğrenebilirsiniz",
        )
        _tg_days = st.selectbox(
            "Kaç iş günü öncesinde bildir?",
            options=[0, 1, 2, 3],
            index=[0,1,2,3].index(int(_tg_cfg.get("notify_days_before", 1))),
            format_func=lambda x: "Sadece rebalance günü" if x == 0 else f"{x} iş günü önce",
            key="tg_days",
        )

        _tg_col1, _tg_col2 = st.columns(2)
        if _tg_col1.button("Kaydet", key="tg_save"):
            _tg_save({
                "bot_token": _tg_token,
                "chat_id": _tg_chat,
                "enabled": _tg_enabled,
                "notify_days_before": _tg_days,
            })
            st.success("Telegram ayarları kaydedildi.")

        if _tg_col2.button("Test Mesajı Gönder", key="tg_test"):
            if not _tg_token or not _tg_chat:
                st.error("Lütfen Bot Token ve Chat ID girin.")
            else:
                from datetime import date as _d2
                _test_msg = "\u2705 <b>Super Investor Telegram baglantisi basarili!</b>\n\nRebalance bildirimleri aktif.\n\n<i>Super Investor - Test Mesaji</i>"
                _ok, _err = _tg_send(_tg_token, _tg_chat, _test_msg)
                if _ok:
                    st.success("Test mesajı gönderildi!")
                else:
                    st.error(f"Gönderilemedi: {_err}")

        if _tg_enabled and not _tg_token:
            st.warning("Bildirim etkin ama Bot Token girilmemiş.")

    except Exception as _tg_err:
        st.error(f"Telegram modülü yüklenemedi: {_tg_err}")

    st.markdown("---")

    market = st.radio(
        "Piyasa",
        options=list(SUPPORTED_MARKETS.keys()),
        format_func=lambda x: SUPPORTED_MARKETS[x]["label"],
        horizontal=True,
    )

    bist_segment = "BISTTUM"
    usa_segment = "USA_ALL"
    if market == "BIST":
        bist_segment = st.selectbox(
            "Evren Seçimi",
            options=list(BIST_SEGMENTS.keys()),
            format_func=lambda x: BIST_SEGMENTS[x],
        )
    else:
        usa_segment = st.selectbox(
            "Evren Seçimi",
            options=list(USA_SEGMENTS.keys()),
            format_func=lambda x: USA_SEGMENTS[x],
        )

    st.divider()

    scan_mode = st.selectbox(
        "Tarama Modu",
        options=list(SCAN_MODES.keys()),
        format_func=lambda x: SCAN_MODES[x],
        index=0,
    )
    st.caption(SCAN_MODE_DESCRIPTIONS[scan_mode])

    st.divider()

    preset_options = get_preset_names()
    selected_preset = st.radio(
        "Temel Kalite Seviyesi",
        options=preset_options,
        format_func=lambda x: QUALITY_LABELS.get(x, x),
        index=0,
        horizontal=True,
    )
    preset_info = get_preset_info(selected_preset)
    if selected_preset != "none" and preset_info["rules"]:
        with st.expander("Filtre kuralları", expanded=False):
            for rule_key, threshold in preset_info["rules"].items():
                st.text(_fmt_rule(rule_key, threshold))

    st.divider()

    sort_by = st.selectbox(
        "Sıralama Türü",
        options=list(SORT_OPTIONS.keys()),
        format_func=lambda x: SORT_OPTIONS[x],
        index=2,
    )

    st.divider()

    inst_profile = st.selectbox(
        "Skor Modeli",
        options=list(STRATEGY_PROFILES.keys()),
        format_func=lambda x: STRATEGY_PROFILES[x]["label"],
        index=0,
    )
    st.caption(STRATEGY_PROFILES[inst_profile]["description"])

    min_avg_volume = st.number_input(
        "Minimum Hacim",
        min_value=0,
        value=0,
        step=100000,
        help="Devre dışı bırakmak için 0",
    )
    min_avg_volume = min_avg_volume if min_avg_volume > 0 else None

    top_n_options = [10, 20, 50, 100]
    top_n = st.selectbox(
        "Sonuç Sayısı",
        options=top_n_options,
        index=1,
    )

    st.divider()

    use_historical = st.checkbox("Geçmiş Tarihte Tara", value=False)
    if use_historical:
        scan_date = st.date_input(
            "Tarama Tarihi",
            value=date.today() - timedelta(days=30),
            min_value=date(2020, 1, 1),
            max_value=date.today() - timedelta(days=1),
        )
        st.caption("Cache'deki fiyat verileri seçilen tarihe kadar kesilir.")
    else:
        # Piyasa kapalı/seans öncesiyse son iş gününü otomatik cutoff olarak kullan
        import datetime as _dt
        _today = date.today()
        _now_hour = _dt.datetime.now().hour  # Yerel saat
        _offset = 0
        _wd = _today.weekday()
        if _wd == 5:    # Cumartesi → Cuma
            _offset = 1
        elif _wd == 6:  # Pazar → Cuma
            _offset = 2
        elif _wd == 0 and _now_hour < 10:   # Pazartesi seans öncesi (BIST 10:00)
            _offset = 3  # Cuma'ya git
        elif _wd in (1,2,3,4) and _now_hour < 10:  # Salı-Cuma seans öncesi
            _offset = 1  # Bir önceki güne git
        _market_closed = _offset > 0
        scan_date = (_today - timedelta(days=_offset)) if _market_closed else None

    st.divider()
    api_status = "Canlı Veri" if TWELVE_DATA_API_KEY else "Demo Veri"
    st.caption(f"Veri: {api_status}")

    with st.expander("Hakkında", expanded=False):
        about_tab1, about_tab2 = st.tabs(["Genel Bilgi", "Doğru Kullanım Şekli"])

        with about_tab1:
            st.markdown(
                """
**RS Skoru** hisseleri beş boyutta 0–100 arası puanlar:

- **Finansal Güç** — ROIC, kaldıraç, varlık kalitesi
- **Büyüme** — Gelir ve kâr büyümesi (YoY + 3Y YBBO)
- **Marj Kalitesi** — Brüt, faaliyet, net, FAVÖK marjları + trend
- **Değerleme** — F/K, PD/DD, FD/FAVÖK, PEG
- **Momentum** — Çoklu dönem getirileri, 52H zirve, göreceli güç
                """
            )
            st.markdown("**Mevcut Piyasalar**")
            for key, info in SUPPORTED_MARKETS.items():
                bench = BENCHMARK_INDEX.get(key, "—")
                st.markdown(f"- **{info['label']}** — {len(info['symbols'])} hisse, endeks: {bench}")

            st.markdown(
                """
**Kalite Filtreleri** sıralama öncesi hisseleri ön eler:
- **Filtre Yok** — tüm evren
- **Temel** — kârlı, pozitif özkaynak, makul kaldıraç
- **Sıkı** — yüksek ROIC, güçlü marjlar, pozitif momentum
                """
            )

tab_screener, tab_backtest, tab_history, tab_guide = st.tabs(["Hisse Tarama", "Backtest", "Geçmiş", "Doğru Kullanım Şekli"])

with tab_screener:
    watchlist_count = len(get_watchlist_tickers())

    run_screening = st.button("Taramayı Başlat", type="primary", use_container_width=True)

    if run_screening:
        market_info = SUPPORTED_MARKETS[market]
        skip_fund = False
        is_historical = use_historical  # Sadece checkbox işaretliyse gerçek geçmiş mod

        try:
            if is_historical:
                cutoff = pd.Timestamp(scan_date)
                with st.spinner(f"{market_info['label']} geçmiş verileri yükleniyor ({scan_date})..."):
                    raw_data, _prep_stats = fetch_backtest_data(
                        market, skip_fundamentals=skip_fund,
                    )
                if raw_data.empty:
                    st.error("Cache'de veri bulunamadı. Önce güncel tarama yaparak verileri indirin.")
                    st.stop()

                from momentum_metrics import append_momentum_fields
                from data_fetcher import get_cached_benchmark
                truncated_rows = []
                from disk_cache import read_cache
                from symbol_mapper import resolve_twelve_symbol
                for _, row in raw_data.iterrows():
                    pd_data = row.get("price_data")
                    if isinstance(pd_data, pd.DataFrame) and "datetime" in pd_data.columns:
                        sliced = pd_data[pd_data["datetime"] <= cutoff].copy()
                        if len(sliced) >= 20:
                            new_row = row.copy()
                            new_row["price_data"] = sliced.reset_index(drop=True)
                            # Tarama tarihindeki fiyatı ham cache'den al (spike filtresi bypass)
                            try:
                                ticker_key = row.get("ticker", "")
                                raw_cache = read_cache(resolve_twelve_symbol(ticker_key, market))
                                if raw_cache is not None and not raw_cache.empty:
                                    raw_cache["datetime"] = pd.to_datetime(raw_cache["datetime"])
                                    raw_sliced = raw_cache[raw_cache["datetime"] <= cutoff]
                                    if not raw_sliced.empty:
                                        price_col = "close" if "close" in raw_sliced.columns else None
                                        if price_col:
                                            last_close = raw_sliced[price_col].iloc[-1]
                                            if pd.notna(last_close):
                                                new_row["price"] = round(float(last_close), 2)
                            except Exception:
                                pass
                            truncated_rows.append(new_row)
                if not truncated_rows:
                    st.error(f"{scan_date} tarihinde yeterli veri bulunamadı.")
                    st.stop()
                raw_data = pd.DataFrame(truncated_rows).reset_index(drop=True)

                bench = get_cached_benchmark(market, cache_only=True)
                if bench is not None and not bench.empty and "datetime" in bench.columns:
                    bench = bench[bench["datetime"] <= cutoff].copy()
                raw_data = append_momentum_fields(raw_data, benchmark_history=bench)
                st.session_state["last_cache_stats"] = None
            else:
                import datetime as _dt2
                _today_wd = date.today().weekday()
                _now_h = _dt2.datetime.now().hour
                _use_cache_only = (
                    _today_wd >= 5 or                          # hafta sonu
                    (_today_wd == 0 and _now_h < 10) or        # Pazartesi seans öncesi
                    (_today_wd in (1,2,3,4) and _now_h < 10)   # Salı-Cuma seans öncesi
                )
                if _use_cache_only:
                    with st.spinner(f"{market_info['label']} verileri işleniyor (seans öncesi — önceki kapanış kullanılıyor)..."):
                        cache_bust = int(time.time() // CACHE_TTL_MARKET_DATA)
                        raw_data = _cached_fetch(market, cache_bust, skip_fundamentals=skip_fund)
                    st.session_state["last_cache_stats"] = None
                else:
                    with st.spinner(f"{market_info['label']} EOD verileri güncelleniyor..."):
                        cache_stats = refresh_eod_cache(market)
                        st.session_state["last_cache_stats"] = cache_stats

                    with st.spinner(f"{market_info['label']} verileri işleniyor..."):
                        cache_bust = int(time.time() // CACHE_TTL_MARKET_DATA)
                        raw_data = _cached_fetch(market, cache_bust, skip_fundamentals=skip_fund)

            if market == "BIST" and bist_segment != "BISTTUM" and "ticker" in raw_data.columns:
                if bist_segment == "BIST100":
                    raw_data = raw_data[raw_data["ticker"].isin(BIST100_TICKERS)].reset_index(drop=True)
                elif bist_segment == "BIST100_DISI":
                    raw_data = raw_data[~raw_data["ticker"].isin(BIST100_TICKERS)].reset_index(drop=True)

            if market == "USA" and usa_segment != "USA_ALL" and "ticker" in raw_data.columns:
                if usa_segment == "SP500":
                    raw_data = raw_data[raw_data["ticker"].isin(SP500_TICKERS)].reset_index(drop=True)
                elif usa_segment == "MIDCAP400":
                    raw_data = raw_data[raw_data["ticker"].isin(MIDCAP400_TICKERS)].reset_index(drop=True)

            validation = validate_dataframe(raw_data)
            if not validation["valid"]:
                st.warning(f"Veri kalitesi uyarısı: eksik sütunlar {validation['missing_columns']}")

            with st.spinner("Puanlama yapılıyor..."):
                scored_data = compute_rs_scores(raw_data, market=market, inst_profile=inst_profile)

            st.session_state["last_scored_data"] = scored_data

            scored_rows = scored_data.to_dict("records")
            updated = update_watchlist_scores(scored_rows)
            if updated > 0:
                st.toast(f"{updated} izleme listesi hissesinin skorları güncellendi")

            filtered_data = apply_preset_filter(
                scored_data, preset=selected_preset,
                min_avg_volume=min_avg_volume, market=market,
            )

            if scan_mode == "smart_money" and not filtered_data.empty:
                mask = pd.Series(True, index=filtered_data.index)
                if "mfi" in filtered_data.columns:
                    mask &= filtered_data["mfi"].fillna(0) > 55
                if "obv_trend_positive" in filtered_data.columns:
                    mask &= filtered_data["obv_trend_positive"].fillna(False)
                if "relative_return_vs_index" in filtered_data.columns:
                    mask &= filtered_data["relative_return_vs_index"].fillna(-999) > 0
                if "volume_ratio" in filtered_data.columns:
                    mask &= filtered_data["volume_ratio"].fillna(0) > 1.10
                if "technical_score" in filtered_data.columns:
                    mask &= filtered_data["technical_score"].fillna(0) >= 55
                if "rs_score" in filtered_data.columns:
                    mask &= filtered_data["rs_score"].fillna(0) >= 60
                filtered_data = filtered_data[mask].reset_index(drop=True)

            elif scan_mode == "early_accumulation" and not filtered_data.empty:
                mask = pd.Series(True, index=filtered_data.index)
                if "mfi" in filtered_data.columns:
                    mfi_col = filtered_data["mfi"].fillna(0)
                    mask &= (mfi_col >= 50) & (mfi_col <= 65)
                if "obv_trend_positive" in filtered_data.columns:
                    mask &= filtered_data["obv_trend_positive"].fillna(False)
                if "distance_to_52w_high" in filtered_data.columns:
                    d52 = filtered_data["distance_to_52w_high"].fillna(-999)
                    mask &= (d52 >= -35) & (d52 <= -10)
                if "return_1m" in filtered_data.columns:
                    mask &= filtered_data["return_1m"].fillna(-999) > -3
                if "return_3m" in filtered_data.columns:
                    mask &= filtered_data["return_3m"].fillna(-999) > 0
                if "rs_score" in filtered_data.columns:
                    mask &= filtered_data["rs_score"].fillna(0) >= 55
                filtered_data = filtered_data[mask].reset_index(drop=True)

            passed_count = len(filtered_data)

            if inst_profile != "standard":
                effective_sort = "institutional_score"
            elif scan_mode != "standard":
                effective_sort = "combined_score"
            else:
                effective_sort = sort_by
            filtered_data = rank_and_limit(filtered_data, top_n=top_n, sort_by=effective_sort)

            st.session_state["screener_scored"] = scored_data
            st.session_state["screener_filtered"] = filtered_data
            st.session_state["screener_passed_count"] = passed_count
            st.session_state["screener_market"] = market
            st.session_state["screener_preset"] = selected_preset
            st.session_state["screener_bist_segment"] = bist_segment if market == "BIST" else None
            st.session_state["screener_usa_segment"] = usa_segment if market == "USA" else None
            st.session_state["screener_sort_by"] = effective_sort
            st.session_state["screener_scan_mode"] = scan_mode
            st.session_state["screener_scan_date"] = scan_date

            segment_label = ""
            if market == "BIST":
                segment_label = BIST_SEGMENTS.get(bist_segment, bist_segment)
            elif market == "USA":
                segment_label = USA_SEGMENTS.get(usa_segment, usa_segment)
            top_tickers = []
            if "ticker" in filtered_data.columns:
                top_tickers = filtered_data["ticker"].head(10).tolist()
            add_scan_entry(
                market=SUPPORTED_MARKETS[market]["label"],
                segment=segment_label,
                scan_mode=SCAN_MODES.get(scan_mode, scan_mode),
                profile=STRATEGY_PROFILES[inst_profile]["label"],
                quality=QUALITY_LABELS.get(selected_preset, selected_preset),
                sort_by=SORT_OPTIONS.get(effective_sort, effective_sort),
                top_n=top_n,
                result_count=passed_count,
                top_stocks=top_tickers,
                scan_date=str(scan_date) if scan_date else None,
            )

        except Exception as e:
            st.error(f"Tarama sırasında bir hata oluştu: {e}")
            _render_diagnostics(None)

    if "screener_filtered" in st.session_state:
        scored_data = st.session_state["screener_scored"]
        filtered_data = st.session_state["screener_filtered"]
        passed_count = st.session_state["screener_passed_count"]
        stored_market = st.session_state["screener_market"]
        stored_preset = st.session_state["screener_preset"]

        stored_scan_date = st.session_state.get("screener_scan_date")
        if stored_scan_date:
            st.info(f"📅 Geçmiş tarih taraması: **{stored_scan_date}** (cache verileri kullanıldı)")

        stored_segment = st.session_state.get("screener_bist_segment")
        stored_usa_segment = st.session_state.get("screener_usa_segment")
        if stored_segment and stored_segment in BIST_SEGMENTS:
            st.caption(f"Evren: **{BIST_SEGMENTS[stored_segment]}**")
        elif stored_usa_segment and stored_usa_segment in USA_SEGMENTS:
            st.caption(f"Evren: **{USA_SEGMENTS[stored_usa_segment]}**")

        stored_scan_mode = st.session_state.get("screener_scan_mode", "standard")
        if stored_scan_mode != "standard":
            mode_label = SCAN_MODES.get(stored_scan_mode, stored_scan_mode)
            st.info(f"Tarama Modu: **{mode_label}**")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Taranan Hisse", len(scored_data))
        c2.metric("Filtreyi Geçen", passed_count)
        if not filtered_data.empty:
            c3.metric("En Yüksek RS", f"{filtered_data['rs_score'].max():.1f}")
            c4.metric("Ort RS", f"{filtered_data['rs_score'].mean():.1f}")
        else:
            c3.metric("En Yüksek RS", "—")
            c4.metric("Ort RS", "—")

        _render_diagnostics(scored_data)

        st.divider()

        # --- Öne Çıkan 5 Hisse: 3 Sütunlu Liste ---
        if not filtered_data.empty:
            def _top5_table(score_col: str, label: str) -> None:
                if score_col not in filtered_data.columns:
                    st.caption(f"{label} verisi yok")
                    return
                _df = (
                    filtered_data.nlargest(5, score_col)[["ticker", "price", score_col]]
                    .reset_index(drop=True)
                )
                _df["#"] = range(1, len(_df) + 1)
                _df["price"] = _df["price"].apply(lambda x: round(x, 2) if not is_na(x) else None)
                _df[score_col] = _df[score_col].apply(lambda x: round(x, 1) if not is_na(x) else None)
                st.dataframe(
                    _df[["#", "ticker", "price"]],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "#": st.column_config.NumberColumn("#", width="small"),
                        "ticker": st.column_config.TextColumn("Hisse"),
                        "price": st.column_config.NumberColumn("Fiyat", format="%.2f"),
                    },
                )

            st.subheader("Öne Çıkan 5 Hisse")
            _col1, _col2, _col3 = st.columns(3)
            with _col1:
                st.markdown("**COMBİNE SKOR**")
                _top5_table("combined_score", "Kombine Skor")
            with _col2:
                st.markdown("**RS SKOR**")
                _top5_table("rs_score", "RS Skor")
            with _col3:
                st.markdown("**TEKNİK SKOR**")
                _top5_table("technical_score", "Teknik Skor")
            st.divider()

        if market == "USA":
            st.caption(
                "Değerleme skoru (F/K, F/DD, EV/FAVÖK, PEG) "
                "sektör içi görece sıralamaya göre hesaplanmaktadır."
            )

        if filtered_data.empty:
            st.warning("Mevcut filtrelere uyan hisse bulunamadı. Kalite filtresini gevşetmeyi veya hacim eşiğini düşürmeyi deneyin.")
        else:
            display_cols = [
                "ticker", "sector", "price",
                "rs_score", "technical_score", "combined_score",
                "institutional_score", "selection_score", "timing_score",
                "inst_category",
                "setup_label", "rs_category",
                "financial_strength", "growth", "margin_quality",
                "valuation", "momentum",
                "inst_quality", "inst_growth", "inst_valuation",
                "inst_momentum", "inst_flow",
                "data_source",
            ]
            display_df = filtered_data[[c for c in display_cols if c in filtered_data.columns]].copy()

            for col in ["rs_score", "technical_score", "combined_score",
                        "institutional_score", "selection_score", "timing_score",
                        "financial_strength", "growth", "margin_quality", "valuation", "momentum",
                        "inst_quality", "inst_growth", "inst_valuation", "inst_momentum", "inst_flow"]:
                if col in display_df.columns:
                    display_df[col] = display_df[col].apply(
                        lambda x: round(x, 1) if not is_na(x) else None
                    )
            if "price" in display_df.columns:
                display_df["price"] = display_df["price"].apply(
                    lambda x: round(x, 2) if not is_na(x) else None
                )

            # Piyasa başlığı
            if market == "BIST":
                st.markdown(
                    "<h3 style='color:#f59e0b;border-bottom:2px solid #f59e0b;padding-bottom:0.4rem;margin-bottom:1rem;'>"
                    f"🇹🇷 BIST Tarama Sonuçları — {len(display_df)} hisse</h3>",
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    "<h3 style='color:#3b82f6;border-bottom:2px solid #3b82f6;padding-bottom:0.4rem;margin-bottom:1rem;'>"
                    f"🇺🇸 USA Tarama Sonuçları — {len(display_df)} hisse</h3>",
                    unsafe_allow_html=True
                )

            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "ticker": st.column_config.TextColumn("Hisse", width="small"),
                    "sector": st.column_config.TextColumn("Sektör"),
                    "price": st.column_config.NumberColumn("Fiyat", format="%.2f"),
                    "rs_score": st.column_config.ProgressColumn(
                        "RS Skoru", min_value=0, max_value=100, format="%.1f",
                    ),
                    "technical_score": st.column_config.ProgressColumn(
                        "Teknik Skor", min_value=0, max_value=100, format="%.1f",
                    ),
                    "combined_score": st.column_config.ProgressColumn(
                        "Kombine Skor", min_value=0, max_value=100, format="%.1f",
                    ),
                    "setup_label": st.column_config.TextColumn("Kurulum", width="small"),
                    "rs_category": st.column_config.TextColumn("Kategori", width="small"),
                    "financial_strength": st.column_config.ProgressColumn(
                        "Finansal", min_value=0, max_value=100, format="%.1f",
                    ),
                    "growth": st.column_config.ProgressColumn(
                        "Büyüme", min_value=0, max_value=100, format="%.1f",
                    ),
                    "margin_quality": st.column_config.ProgressColumn(
                        "Marjlar", min_value=0, max_value=100, format="%.1f",
                    ),
                    "valuation": st.column_config.ProgressColumn(
                        "Değerleme", min_value=0, max_value=100, format="%.1f",
                    ),
                    "momentum": st.column_config.ProgressColumn(
                        "Momentum", min_value=0, max_value=100, format="%.1f",
                    ),
                    "institutional_score": st.column_config.ProgressColumn(
                        "Kurumsal Skor", min_value=0, max_value=100, format="%.1f",
                    ),
                    "selection_score": st.column_config.ProgressColumn(
                        "Seçim Sk.", min_value=0, max_value=100, format="%.1f",
                    ),
                    "timing_score": st.column_config.ProgressColumn(
                        "Zamanlama Sk.", min_value=0, max_value=100, format="%.1f",
                    ),
                    "inst_category": st.column_config.TextColumn("Kurumsal Kat.", width="small"),
                    "inst_quality": st.column_config.ProgressColumn(
                        "K.Kalite", min_value=0, max_value=100, format="%.1f",
                    ),
                    "inst_growth": st.column_config.ProgressColumn(
                        "K.Büyüme", min_value=0, max_value=100, format="%.1f",
                    ),
                    "inst_valuation": st.column_config.ProgressColumn(
                        "K.Değerleme", min_value=0, max_value=100, format="%.1f",
                    ),
                    "inst_momentum": st.column_config.ProgressColumn(
                        "K.Momentum", min_value=0, max_value=100, format="%.1f",
                    ),
                    "inst_flow": st.column_config.ProgressColumn(
                        "K.Akış", min_value=0, max_value=100, format="%.1f",
                    ),
                    "data_source": st.column_config.TextColumn("Kaynak", width="small"),
                },
            )

            csv_cols = [
                "rank", "ticker", "company_name", "sector", "price", "market_cap",
                "rs_score", "technical_score", "combined_score",
                "institutional_score", "selection_score", "timing_score", "inst_category",
                "setup_label", "rs_category",
                "financial_strength", "growth", "margin_quality", "valuation", "momentum",
                "inst_quality", "inst_growth", "inst_valuation", "inst_momentum", "inst_flow",
                "return_1m", "return_3m", "return_6m", "return_12m",
                "data_source",
            ]
            csv_df = filtered_data[[c for c in csv_cols if c in filtered_data.columns]].copy()
            csv_data = csv_df.to_csv(index=False)

            st.download_button(
                label="CSV İndir",
                data=csv_data,
                file_name=f"tarama_{stored_market}_{stored_preset}.csv",
                mime="text/csv",
            )

            st.divider()
            st.subheader("Hisse Detayları")

            for _, row in filtered_data.iterrows():
                ticker = row.get("ticker", "?")
                name = row.get("company_name", "")
                cat = row.get("rs_category", "N/A")
                rs = row.get("rs_score", 0)
                setup = row.get("setup_label", "")
                in_wl = is_in_watchlist(ticker)
                wl_icon = " [İ]" if in_wl else ""
                setup_tag = f"  · {setup}" if setup and setup != "N/A" else ""
                label = f"{ticker}  —  {name}  |  RS {rs:.1f}  ({cat}){setup_tag}{wl_icon}"

                with st.expander(label, expanded=False):
                    if not in_wl:
                        if st.button(f"{ticker} İzleme Listesine Ekle", key=f"wl_add_{ticker}"):
                            ok = add_to_watchlist(
                                ticker=ticker,
                                rs_score=rs,
                                rs_category=cat,
                                price=row.get("price"),
                                company_name=name,
                                market=stored_market,
                            )
                            if ok:
                                st.success(f"{ticker} izleme listesine eklendi")
                            else:
                                st.error(f"{ticker} izleme listesine kaydedilemedi")
                            st.rerun()
                    else:
                        st.caption(f"{ticker} izleme listenizde")
                        if st.button(f"{ticker} İzleme Listesinden Kaldır", key=f"wl_rem_{ticker}"):
                            remove_from_watchlist(ticker)
                            st.rerun()

                    _render_detail(row, scored_df=st.session_state.get("last_scored_data"))

    else:
        st.markdown("Bir piyasa seçin ve hisseleri skorlarına göre sıralamak için **Taramayı Başlat** butonuna tıklayın.")
        st.info("Daha fazla bilgi için **Doğru Kullanım Şekli** sekmesine göz atabilirsiniz.")

    st.divider()
    wl_items = get_watchlist()
    wl_count = len(wl_items)
    with st.expander(f"İzleme Listesi ({wl_count})", expanded=False):
        if not wl_items:
            st.info("İzleme listeniz boş. Taramayı çalıştırın ve hisse ekleyin.")
        else:
            wl_df = pd.DataFrame(wl_items)
            display_wl_cols = ["ticker", "company_name", "rs_score", "rs_category", "price", "market"]
            display_wl = wl_df[[c for c in display_wl_cols if c in wl_df.columns]].copy()

            if "rs_score" in display_wl.columns:
                display_wl["rs_score"] = display_wl["rs_score"].apply(
                    lambda x: round(x, 1) if not is_na(x) else None
                )
            if "price" in display_wl.columns:
                display_wl["price"] = display_wl["price"].apply(
                    lambda x: round(x, 2) if not is_na(x) else None
                )

            st.dataframe(
                display_wl,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "ticker": st.column_config.TextColumn("Hisse", width="small"),
                    "company_name": st.column_config.TextColumn("Şirket"),
                    "rs_score": st.column_config.ProgressColumn(
                        "RS Skoru", min_value=0, max_value=100, format="%.1f",
                    ),
                    "rs_category": st.column_config.TextColumn("Kategori", width="small"),
                    "price": st.column_config.NumberColumn("Fiyat", format="%.2f"),
                    "market": st.column_config.TextColumn("Piyasa", width="small"),
                },
            )

            col_csv, col_clear = st.columns([1, 1])
            with col_csv:
                csv_wl = export_watchlist_csv()
                st.download_button(
                    label="İzleme Listesini CSV İndir",
                    data=csv_wl,
                    file_name="izleme_listesi.csv",
                    mime="text/csv",
                )
            with col_clear:
                if st.button("Listeyi Temizle", type="secondary"):
                    cleared = clear_watchlist()
                    st.success(f"{cleared} hisse izleme listesinden kaldırıldı.")
                    st.rerun()

            for item in wl_items:
                t = item["ticker"]
                col_t, col_btn = st.columns([3, 1])
                col_t.text(f"{t} — {item.get('company_name', '')} | RS {_score_fmt(item.get('rs_score'))}")
                if col_btn.button("Kaldır", key=f"wl_remove_{t}"):
                    remove_from_watchlist(t)
                    st.rerun()

with tab_backtest:
    from backtest_engine import run_backtest
    from datetime import date, timedelta

    st.markdown("Tarama stratejisinin geçmiş performansını simüle edin.")
    st.info("Backtest sadece disk cache'indeki verileri kullanır. Verilerin güncel olması için önce **Hisse Tarama** sekmesinden tarama yapın.")

    bt_p1, bt_p2, bt_p3 = st.columns(3)

    with bt_p1:
        bt_market = st.radio(
            "Piyasa",
            options=list(SUPPORTED_MARKETS.keys()),
            format_func=lambda x: SUPPORTED_MARKETS[x]["label"],
            horizontal=True,
            key="bt_market",
        )

        bt_bist_segment = "BISTTUM"
        bt_usa_segment = "USA_ALL"
        if bt_market == "BIST":
            bt_bist_segment = st.selectbox(
                "Evren Seçimi",
                options=list(BIST_SEGMENTS.keys()),
                format_func=lambda x: BIST_SEGMENTS[x],
                key="bt_universe_bist",
            )
        else:
            bt_usa_segment = st.selectbox(
                "Evren Seçimi",
                options=list(USA_SEGMENTS.keys()),
                format_func=lambda x: USA_SEGMENTS[x],
                key="bt_universe_usa",
            )

        bt_profile = st.selectbox(
            "Strateji Profili",
            options=list(STRATEGY_PROFILES.keys()),
            format_func=lambda x: STRATEGY_PROFILES[x]["label"],
            index=0,
            key="bt_profile",
        )
        st.caption(STRATEGY_PROFILES[bt_profile]["description"])

        bt_scan_mode = st.selectbox(
            "Ek Filtre (Opsiyonel)",
            options=list(SCAN_MODES.keys()),
            format_func=lambda x: SCAN_MODES[x],
            index=0,
            key="bt_scan_mode",
        )

    with bt_p2:
        bt_preset = st.radio(
            "Temel Kalite Seviyesi",
            options=get_preset_names(),
            format_func=lambda x: QUALITY_LABELS.get(x, x),
            index=0,
            horizontal=True,
            key="bt_preset",
        )

        bt_sort = st.selectbox(
            "Sıralama Türü",
            options=list(SORT_OPTIONS.keys()),
            format_func=lambda x: SORT_OPTIONS[x],
            index=2,
            key="bt_sort",
        )

        bt_top_n = st.selectbox(
            "Portföy Hisse Sayısı",
            options=[3, 5, 7, 9],
            index=1,
            key="bt_top_n",
        )

        bt_rebalance = st.selectbox(
            "Rebalance Sıklığı",
            options=list(REBALANCE_LABELS.keys()),
            format_func=lambda x: REBALANCE_LABELS[x],
            index=2,
            key="bt_rebalance",
        )

    with bt_p3:
        bt_start = st.date_input(
            "Başlangıç Tarihi",
            value=date.today() - timedelta(days=180),
            max_value=date.today() - timedelta(days=30),
            key="bt_start",
        )

        bt_end = st.date_input(
            "Bitiş Tarihi",
            value=date.today(),
            max_value=date.today(),
            key="bt_end",
        )

        bt_benchmark_label = BENCHMARK_INDEX.get(bt_market, "—")
        st.markdown(f"**Benchmark:** {bt_benchmark_label}")

        bt_weight = st.selectbox(
            "Ağırlıklandırma",
            options=list(WEIGHT_LABELS.keys()),
            format_func=lambda x: WEIGHT_LABELS[x],
            index=0,
            key="bt_weight",
        )

    if bt_start >= bt_end:
        st.warning("Başlangıç tarihi bitiş tarihinden önce olmalıdır.")
    else:
        run_bt = st.button("Backtest Başlat", type="primary", use_container_width=True, key="bt_run")

        if run_bt:
            bt_universe = bt_bist_segment if bt_market == "BIST" else bt_usa_segment

            progress_bar = st.progress(0, text="Backtest başlatılıyor...")

            def _bt_progress(pct, text):
                progress_bar.progress(min(pct, 1.0), text=text)

            try:
                with st.spinner("Veriler çekilip analiz ediliyor..."):
                    result = run_backtest(
                        market=bt_market,
                        universe=bt_universe,
                        scan_mode=bt_scan_mode,
                        quality_preset=bt_preset,
                        sort_by=bt_sort,
                        top_n=bt_top_n,
                        rebalance_freq=bt_rebalance,
                        start_date=bt_start,
                        end_date=bt_end,
                        progress_callback=_bt_progress,
                        inst_profile=bt_profile,
                    )

                progress_bar.empty()

                bt_universe_label = (
                    BIST_SEGMENTS.get(bt_bist_segment, bt_bist_segment)
                    if bt_market == "BIST"
                    else USA_SEGMENTS.get(bt_usa_segment, bt_usa_segment)
                )

                st.session_state["bt_result"] = result
                if bt_profile != "standard":
                    effective_sort = "Institutional Score"
                elif bt_scan_mode != "standard":
                    effective_sort = "Combined Score"
                else:
                    effective_sort = SORT_OPTIONS.get(bt_sort, bt_sort)

                bt_params_dict = {
                    "market": SUPPORTED_MARKETS[bt_market]["label"],
                    "universe": bt_universe_label,
                    "profile": STRATEGY_PROFILES[bt_profile]["label"],
                    "scan_mode": SCAN_MODES.get(bt_scan_mode, bt_scan_mode),
                    "quality": QUALITY_LABELS.get(bt_preset, bt_preset),
                    "sort_by": effective_sort,
                    "top_n": bt_top_n,
                    "rebalance": REBALANCE_LABELS.get(bt_rebalance, bt_rebalance),
                    "weight": WEIGHT_LABELS.get(bt_weight, bt_weight),
                    "benchmark": bt_benchmark_label,
                    "start": str(bt_start),
                    "end": str(bt_end),
                }
                st.session_state["bt_params"] = bt_params_dict

                add_backtest_entry(
                    params=bt_params_dict,
                    total_return=result.total_return,
                    benchmark_return=result.benchmark_return,
                    num_periods=result.num_periods,
                    sharpe=result.sharpe_ratio,
                    max_drawdown=result.max_drawdown,
                    rebalance_history=result.rebalance_history,
                )

            except Exception as e:
                progress_bar.empty()
                st.error(f"Backtest sırasında hata oluştu: {e}")

    if "bt_result" in st.session_state:
        result = st.session_state["bt_result"]
        params = st.session_state.get("bt_params", {})

        if result.num_periods == 0:
            st.warning("Seçilen tarih aralığında yeterli veri bulunamadı. Tarih aralığını genişletmeyi deneyin.")
        else:
            st.divider()

            if result.data_prep_stats:
                dps = result.data_prep_stats
                prep_parts = [f"Toplam: {dps.total_symbols}"]
                if dps.cache_hits > 0:
                    prep_parts.append(f"Cache: {dps.cache_hits}")
                if dps.incremental_updates > 0:
                    prep_parts.append(f"Guncellenen: {dps.incremental_updates}")
                if dps.full_fetches > 0:
                    prep_parts.append(f"Yeni cekilen: {dps.full_fetches}")
                if dps.failed > 0:
                    prep_parts.append(f"Basarisiz: {dps.failed}")
                if dps.used_mock:
                    prep_parts.append("Demo veri")
                prep_parts.append(f"Sure: {dps.duration_seconds:.1f}s")
                st.caption("Veri Durumu: " + " · ".join(prep_parts))
                if dps.failed_symbols:
                    st.warning(f"Basarisiz semboller: {', '.join(dps.failed_symbols[:20])}")

            info_col1, info_col2 = st.columns(2)
            with info_col1:
                st.markdown(f"**Baslangic:** {params.get('start', '—')}  ·  **Bitis:** {params.get('end', '—')}")
                st.markdown(f"**Periyot:** {params.get('rebalance', '—')}  ·  **Donem Sayisi:** {result.num_periods}")
            with info_col2:
                cond_parts = [
                    f"Piyasa: {params.get('market', '?')}",
                    f"Evren: {params.get('universe', '?')}",
                    f"Strateji: {params.get('profile', '?')}",
                    f"Filtre: {params.get('scan_mode', '?')}",
                    f"Kalite: {params.get('quality', '?')}",
                    f"Siralama: {params.get('sort_by', '?')}",
                    f"Top-{params.get('top_n', '?')}",
                    f"Agirlik: {params.get('weight', '?')}",
                    f"Benchmark: {params.get('benchmark', '?')}",
                ]
                st.caption("Kosullar: " + " · ".join(cond_parts))

            st.divider()
            st.markdown("#### Performans Özeti")

            m1, m2, m3 = st.columns(3)
            m1.metric("Portföy Getirisi", f"%{result.total_return:.1f}")
            m2.metric("Benchmark Getirisi", f"%{result.benchmark_return:.1f}")
            m3.metric("Alpha", f"%{result.alpha:.1f}")

            m4, m5, m6 = st.columns(3)
            m4.metric("Maks. Drawdown", f"%{result.max_drawdown:.1f}")
            m5.metric("Sharpe Oranı", f"{result.sharpe_ratio:.2f}")
            m6.metric("Volatilite (Yıllık)", f"%{result.volatility:.1f}")

            st.divider()
            st.markdown("#### Equity Eğrisi")

            eq_df = result.equity_curve.copy()
            bench_df = result.benchmark_curve.copy()

            if not eq_df.empty and not bench_df.empty:
                chart_data = pd.DataFrame({
                    "Tarih": eq_df["date"],
                    "Portföy": eq_df["value"],
                    "Benchmark": bench_df["value"],
                }).set_index("Tarih")
                st.line_chart(chart_data, use_container_width=True)

            st.divider()
            st.markdown("#### Drawdown Grafiği")

            if not result.drawdown_series.empty:
                dd_chart = pd.DataFrame({
                    "Portföy DD (%)": (result.drawdown_series.values * 100),
                    "Benchmark DD (%)": (result.benchmark_drawdown_series.values * 100),
                }, index=result.drawdown_series.index)
                st.area_chart(dd_chart, use_container_width=True)

            st.divider()
            st.markdown("#### Rebalance Geçmişi")

            if result.rebalance_history:
                history_rows = []
                for rec in result.rebalance_history:
                    history_rows.append({
                        "Tarih": rec.date.strftime("%Y-%m-%d"),
                        "Hisse Sayısı": len(rec.tickers),
                        "Seçilen Hisseler": ", ".join(rec.tickers[:10]) + ("..." if len(rec.tickers) > 10 else ""),
                        "Dönem Getirisi (%)": rec.period_return,
                    })
                history_df = pd.DataFrame(history_rows)
                st.dataframe(history_df, use_container_width=True, hide_index=True)

                with st.expander("Detaylı Dönem Skorları", expanded=False):
                    for rec in result.rebalance_history:
                        if rec.tickers:
                            st.markdown(f"**{rec.date.strftime('%Y-%m-%d')}** — Getiri: %{rec.period_return:.1f}")
                            if rec.ticker_returns:
                                return_items = [f"{t}: %{r:+.1f}" for t, r in rec.ticker_returns.items()]
                            else:
                                return_items = [f"{t}: %{s:.1f}" for t, s in rec.scores.items()]
                            st.caption(" · ".join(return_items))

def _build_backtest_excel(entry: dict) -> bytes:
    """
    Backtest kaydını Excel dosyasına dönüştürür.

    Sheet 1 – Özet: Parametreler ve genel istatistikler
    Sheet 2 – İşlem Geçmişi: Rebalance tarihleri + alış/satış/tutulan hisseler
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    params = entry.get("params", {})
    reb_hist = entry.get("rebalance_history", [])

    try: total_ret  = float(entry.get("total_return", 0) or 0)
    except: total_ret = 0.0
    try: bench_ret  = float(entry.get("benchmark_return", 0) or 0)
    except: bench_ret = 0.0
    try: sharpe_val = float(entry.get("sharpe", 0) or 0)
    except: sharpe_val = 0.0
    try: dd_val     = float(entry.get("max_drawdown", 0) or 0)
    except: dd_val = 0.0

    wb = Workbook()

    # ── RENK PALETİ ──────────────────────────────────────
    COLOR_HEADER_BG  = "1E3A5F"   # koyu lacivert
    COLOR_HEADER_FG  = "FFFFFF"   # beyaz
    COLOR_LABEL_BG   = "2D6A9F"   # orta mavi
    COLOR_ALT_ROW    = "EBF3FB"   # açık mavi (zebra)
    COLOR_SECTION    = "2563EB"   # bölüm başlığı
    COLOR_BUY        = "D1FAE5"   # yeşil
    COLOR_SELL       = "FEE2E2"   # kırmızı
    COLOR_HOLD       = "FEF9C3"   # sarı
    COLOR_POS        = "16A34A"   # pozitif getiri yazısı
    COLOR_NEG        = "DC2626"   # negatif getiri yazısı

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, val, bold=True, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, sz=11):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=sz)
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        return c

    def _val(ws, row, col, val, bold=False, bg="FFFFFF", align="left", number_fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=10)
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = border
        if number_fmt:
            c.number_format = number_fmt
        return c

    def _section(ws, row, text):
        ws.row_dimensions[row].height = 22
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color=COLOR_SECTION, size=11)
        c.alignment = Alignment(vertical="center")

    # ══════════════════════════════════════════════════════
    # SHEET 1 – ÖZET
    # ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Özet"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 32

    # Başlık bandı
    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    c.value = "SUPER INVESTOR — BACKTEST RAPORU"
    c.font = Font(bold=True, color=COLOR_HEADER_FG, size=13)
    c.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    r = 2
    # ── BACKTEST PARAMETRELERİ ──
    _section(ws1, r, "BACKTEST PARAMETRELERİ"); r += 1

    params_rows = [
        ("Piyasa",                 params.get("market", "")),
        ("Başlangıç Tarihi",       params.get("start", "")),
        ("Bitiş Tarihi",           params.get("end",   "")),
        ("Portföy Hisse Sayısı",   str(params.get("top_n", ""))),
        ("Rebalance Sıklığı",      params.get("rebalance", "")),
        ("Temel Kalite Seviyesi",  params.get("quality", "")),
        ("Sıralama Türü",          params.get("sort_by", "")),
        ("Strateji Profili",       params.get("profile", "")),
        ("Evren",                  params.get("universe", "")),
        ("Tarama Modu",            params.get("scan_mode", "")),
    ]

    for label, val in params_rows:
        bg = COLOR_ALT_ROW if r % 2 == 0 else "FFFFFF"
        _hdr(ws1, r, 1, label, bg=COLOR_LABEL_BG, sz=10)
        _val(ws1, r, 2, val, bold=True, bg=bg, align="center")
        r += 1

    r += 1
    # ── PERFORMANS SONUÇLARI ──
    _section(ws1, r, "PERFORMANS SONUÇLARI"); r += 1

    perf_rows = [
        ("Toplam Getiri",        f"%{total_ret:.2f}"),
        ("Benchmark Getirisi",   f"%{bench_ret:.2f}"),
        ("Fazla Getiri (Alpha)",  f"%{total_ret - bench_ret:.2f}"),
        ("Sharpe Oranı",         f"{sharpe_val:.3f}"),
        ("Maksimum Drawdown",    f"%{dd_val:.2f}"),
        ("Dönem Sayısı",         str(entry.get("num_periods", ""))),
        ("Backtest Tarihi",      entry.get("datetime", "")),
    ]

    for label, val in perf_rows:
        bg = COLOR_ALT_ROW if r % 2 == 0 else "FFFFFF"
        _hdr(ws1, r, 1, label, bg=COLOR_LABEL_BG, sz=10)
        c = _val(ws1, r, 2, val, bold=True, bg=bg, align="center")
        # Renk: pozitif/negatif getiri
        if "Getiri" in label or "Alpha" in label:
            try:
                num = float(val.replace("%",""))
                c.font = Font(bold=True, color=COLOR_POS if num >= 0 else COLOR_NEG, size=10)
            except: pass
        r += 1

    # ══════════════════════════════════════════════════════
    # SHEET 2 – İŞLEM GEÇMİŞİ
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="İşlem Geçmişi")

    # Sütun genişlikleri
    col_widths = {1: 14, 2: 16, 3: 36, 4: 36, 5: 36, 6: 36}
    for c_idx, w in col_widths.items():
        ws2.column_dimensions[get_column_letter(c_idx)].width = w

    # Başlık bandı
    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    piyasa_str = params.get("market", "")
    tarih_str  = f"{params.get('start','')} – {params.get('end','')}"
    c.value = f"{piyasa_str}  |  {tarih_str}  |  Top {params.get('top_n','')} Hisse  |  {params.get('rebalance','')} Rebalance"
    c.font = Font(bold=True, color=COLOR_HEADER_FG, size=12)
    c.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    # Sütun başlıkları
    headers = ["Tarih", "Dönem Getirisi", "Portföydeki Hisseler",
               "Alınan (Yeni Giriş)", "Satılan (Çıkış)", "Tutulan"]
    for col_i, h in enumerate(headers, 1):
        _hdr(ws2, 2, col_i, h, sz=10)
    ws2.row_dimensions[2].height = 20

    if not reb_hist:
        ws2.merge_cells("A3:F3")
        ws2["A3"].value = "İşlem geçmişi bulunamadı."
    else:
        for r_idx, rec in enumerate(reb_hist):
            row_n  = r_idx + 3
            tickers      = rec.get("tickers", [])
            prev_tickers = reb_hist[r_idx-1].get("tickers", []) if r_idx > 0 else []
            sold   = [t for t in prev_tickers if t not in tickers]
            bought = [t for t in tickers      if t not in prev_tickers]
            held   = [t for t in tickers      if t in  prev_tickers]

            try:
                pret = float(rec.get("period_return", 0) or 0)
            except:
                pret = 0.0

            # Zebralar
            row_bg = COLOR_ALT_ROW if r_idx % 2 == 0 else "FFFFFF"

            # Tarih
            _val(ws2, row_n, 1, str(rec.get("date", "")),
                 bg=row_bg, align="center")

            # Dönem getirisi — renkli
            c_ret = _val(ws2, row_n, 2, f"%{pret:.1f}",
                         bold=True, bg=row_bg, align="center")
            c_ret.font = Font(bold=True, size=10,
                              color=COLOR_POS if pret >= 0 else COLOR_NEG)

            # Portföydeki hisseler
            _val(ws2, row_n, 3, ", ".join(tickers), bg=row_bg)

            # Alınan → yeşil
            c_buy = _val(ws2, row_n, 4,
                         ", ".join(bought) if bought else "—",
                         bg=COLOR_BUY if bought else row_bg)
            if bought:
                c_buy.font = Font(bold=True, color="166534", size=10)

            # Satılan → kırmızı
            c_sell = _val(ws2, row_n, 5,
                          ", ".join(sold) if sold else "—",
                          bg=COLOR_SELL if sold else row_bg)
            if sold:
                c_sell.font = Font(bold=True, color="991B1B", size=10)

            # Tutulan → sarı
            _val(ws2, row_n, 6,
                 ", ".join(held) if held else "—",
                 bg=COLOR_HOLD if held else row_bg)

            ws2.row_dimensions[row_n].height = 18

    # ══════════════════════════════════════════════════════
    # SHEET 3 – HISSE BAZLI GETİRİLER
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet(title="Hisse Bazlı Getiriler")
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 22
    ws3.row_dimensions[1].height = 24

    ws3.merge_cells("A1:C1")
    c = ws3["A1"]
    c.value = "DÖNEM BAZLI HİSSE GETİRİLERİ"
    c.font = Font(bold=True, color=COLOR_HEADER_FG, size=12)
    c.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    _hdr(ws3, 2, 1, "Tarih", sz=10)
    _hdr(ws3, 2, 2, "Hisse", sz=10)
    _hdr(ws3, 2, 3, "Getiri (%)", sz=10)

    r3 = 3
    if reb_hist:
        for rec in reb_hist:
            tr = rec.get("ticker_returns", {})
            if tr:
                for ticker, ret in tr.items():
                    try: ret_f = float(ret)
                    except: ret_f = 0.0
                    bg = COLOR_ALT_ROW if r3 % 2 == 0 else "FFFFFF"
                    _val(ws3, r3, 1, str(rec.get("date","")), bg=bg, align="center")
                    _val(ws3, r3, 2, str(ticker), bg=bg, align="center")
                    c_r = _val(ws3, r3, 3, round(ret_f, 2), bg=bg, align="center", number_fmt="0.00")
                    c_r.font = Font(bold=False, size=10,
                                   color=COLOR_POS if ret_f >= 0 else COLOR_NEG)
                    r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


with tab_history:
    if "pending_delete_id" in st.session_state:
        _del_id = st.session_state.pop("pending_delete_id")
        delete_entry(_del_id)
        st.rerun()
    if "pending_clear_history" in st.session_state:
        st.session_state.pop("pending_clear_history")
        clear_history()
        st.rerun()

    st.subheader("Tarama ve Backtest Geçmişi")
    history_entries = get_history()

    if not history_entries:
        st.info("Henüz kayıtlı tarama veya backtest bulunmuyor.")
    else:
        col_info, col_clear = st.columns([4, 1])
        with col_info:
            scan_count = sum(1 for e in history_entries if e.get("type") == "scan")
            bt_count   = sum(1 for e in history_entries if e.get("type") == "backtest")
            st.caption(f"Toplam {len(history_entries)} kayıt ({scan_count} tarama, {bt_count} backtest)")
        with col_clear:
            st.button("Tüm Geçmişi Temizle", type="secondary", key="clear_all_history",
                      on_click=lambda: st.session_state.update({"pending_clear_history": True}))

        def _market_of(e):
            m = str(e.get("market") or e.get("params", {}).get("market", "") or "")
            m_up = m.upper()
            if "BIST" in m_up or "TÜRKİYE" in m_up or "TURKEY" in m_up:
                return "BIST"
            return "USA"

        bist_bt   = [e for e in history_entries if e.get("type") == "backtest" and _market_of(e) == "BIST"]
        usa_bt    = [e for e in history_entries if e.get("type") == "backtest" and _market_of(e) == "USA"]
        bist_scan = [e for e in history_entries if e.get("type") == "scan"     and _market_of(e) == "BIST"]
        usa_scan  = [e for e in history_entries if e.get("type") == "scan"     and _market_of(e) == "USA"]

        def _render_scan(entries, base_key):
            for idx, entry in enumerate(entries):
                entry_id  = entry.get("id", f"{base_key}_{idx}")
                entry_dt  = entry.get("datetime", "")
                scan_date_str = entry.get("scan_date")
                date_badge    = f" | Tarih: {scan_date_str}" if scan_date_str else ""
                header = f"TARAMA — {entry.get('market','')} | {entry.get('segment','')} | {entry_dt}{date_badge}"
                with st.expander(header, expanded=False):
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Mod",      entry.get("scan_mode", ""))
                    c2.metric("Profil",   entry.get("profile", ""))
                    c3.metric("Kalite",   entry.get("quality", ""))
                    c4.metric("Sıralama", entry.get("sort_by", ""))
                    c5, c6 = st.columns(2)
                    c5.metric("Top N",        entry.get("top_n", ""))
                    c6.metric("Sonuç Sayısı", entry.get("result_count", ""))
                    top_stocks = entry.get("top_stocks", [])
                    if top_stocks:
                        st.markdown(f"**En İyi Hisseler:** {', '.join(top_stocks)}")
                    _scan_id = entry_id
                    st.button("Bu Kaydı Sil", key=f"del_scan_{base_key}_{idx}", type="secondary",
                              on_click=lambda eid=_scan_id: st.session_state.update({"pending_delete_id": eid}))

        def _render_bt(entries, base_key):
            for idx, entry in enumerate(entries):
                entry_id = entry.get("id", f"{base_key}_{idx}")
                entry_dt = entry.get("datetime", "")
                params   = entry.get("params", {})
                try:    total_ret = float(entry.get("total_return", 0) or 0)
                except: total_ret = 0.0
                try:    bench_ret = float(entry.get("benchmark_return", 0) or 0)
                except: bench_ret = 0.0
                try:    sharpe_val = float(entry.get("sharpe", 0) or 0)
                except: sharpe_val = 0.0
                try:    dd_val = float(entry.get("max_drawdown", 0) or 0)
                except: dd_val = 0.0
                ret_sign = "+" if total_ret >= 0 else ""
                _bt_id = entry_id
                header = f"BACKTEST — {params.get('market','')} | {ret_sign}%{total_ret:.1f} | {params.get('start','')} ~ {params.get('end','')} | {entry_dt}"
                _row_col1, _row_col2 = st.columns([11, 1])
                with _row_col2:
                    st.button("🗑️", key=f"del_bt_{base_key}_{idx}", help="Bu kaydı sil",
                              on_click=lambda eid=_bt_id: st.session_state.update({"pending_delete_id": eid}))
                with _row_col1:
                    with st.expander(header, expanded=False):
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("Getiri",       f"%{total_ret:.1f}", delta=f"vs Benchmark %{bench_ret:.1f}")
                        c2.metric("Sharpe",       f"{sharpe_val:.2f}")
                        c3.metric("Max Drawdown", f"%{dd_val:.1f}")
                        c4.metric("Dönem Sayısı", entry.get("num_periods", 0))
                        c5, c6, c7, c8 = st.columns(4)
                        c5.metric("Evren",    params.get("universe", ""))
                        c6.metric("Profil",   params.get("profile", ""))
                        c7.metric("Mod",      params.get("scan_mode", ""))
                        c8.metric("Rebalance",params.get("rebalance", ""))
                        c9, c10, c11 = st.columns(3)
                        c9.metric("Kalite",       params.get("quality", ""))
                        c10.metric("Top N",       params.get("top_n", ""))
                        c11.metric("Sıralama Türü", params.get("sort_by", ""))
                        reb_hist = entry.get("rebalance_history", [])
                        if reb_hist:
                            st.markdown("---")
                            st.markdown("**İşlem Listesi**")
                            _rows = []
                            for _ri, _rec in enumerate(reb_hist):
                                _tickers      = _rec.get("tickers", [])
                                _prev_tickers = reb_hist[_ri-1].get("tickers", []) if _ri > 0 else []
                                _sold    = [t for t in _prev_tickers if t not in _tickers]
                                _bought  = [t for t in _tickers if t not in _prev_tickers]
                                _held    = [t for t in _tickers if t in _prev_tickers]
                                _rows.append({
                                    "Tarih":        _rec.get("date",""),
                                    "Dönem Getirisi": f"%{_rec.get('period_return',0):.1f}",
                                    "Hisseler":     ", ".join(_tickers),
                                    "Alınan":       ", ".join(_bought) if _bought else "-",
                                    "Satılan":      ", ".join(_sold) if _sold else "-",
                                    "Tutulan":      ", ".join(_held) if _held else "-",
                                })
                            st.dataframe(pd.DataFrame(_rows), use_container_width=True, hide_index=True)
                            with st.expander("Detaylı Dönem Skorları", expanded=False):
                                for _rec in reb_hist:
                                    _tickers = _rec.get("tickers", [])
                                    _ret     = _rec.get("period_return", 0)
                                    _tr      = _rec.get("ticker_returns", {})
                                    st.markdown(f"**{_rec.get('date','')}** — Dönem getirisi: **%{_ret:.1f}**")
                                    if _tr:
                                        items = [f"{t}: %{r:+.1f}" for t, r in _tr.items()]
                                    else:
                                        items = _tickers
                                    st.caption(" · ".join(items))

                        # Excel indirme butonu
                        st.markdown("---")
                        _p = entry.get("params", {})
                        _fname = (
                            f"backtest_{_p.get('market','MKT')}"
                            f"_{_p.get('start','')}_{_p.get('end','')}"
                            f"_top{_p.get('top_n','')}.xlsx"
                        ).replace(" ", "_")
                        _downloads_dir = os.path.expanduser("~/Downloads")
                        _save_path = os.path.join(_downloads_dir, _fname)
                        _excel_col1, _excel_col2, _excel_col3 = st.columns([2, 1, 1])
                        try:
                            _xl_bytes = _build_backtest_excel(entry)
                            with _excel_col2:
                                st.download_button(
                                    label="Tarayıcıdan İndir",
                                    data=_xl_bytes,
                                    file_name=_fname,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"xl_{base_key}_{idx}",
                                    use_container_width=True,
                                )
                            with _excel_col3:
                                if st.button("Downloads Klasörüne Kaydet", key=f"xl_save_{base_key}_{idx}",
                                             use_container_width=True):
                                    os.makedirs(_downloads_dir, exist_ok=True)
                                    with open(_save_path, "wb") as _fh:
                                        _fh.write(_xl_bytes)
                                    st.success(f"Kaydedildi: ~/Downloads/{_fname}")
                        except Exception as _xl_err:
                            st.warning(f"Excel oluşturulamadı: {_xl_err}")

        # ── BACKTEST SONUÇLARI ──────────────────────────────────────────────
        if bist_bt or usa_bt:
            st.markdown("---")
            st.markdown("### 📈 Backtest Sonuçları")
            if bist_bt:
                st.markdown(
                    "<h4 style='color:#f59e0b;border-left:4px solid #f59e0b;padding-left:0.6rem;'>🇹🇷 BIST Backtest Sonuçları</h4>",
                    unsafe_allow_html=True)
                _render_bt(bist_bt, "bist_bt")
            if usa_bt:
                st.markdown(
                    "<h4 style='color:#3b82f6;border-left:4px solid #3b82f6;padding-left:0.6rem;'>🇺🇸 USA Backtest Sonuçları</h4>",
                    unsafe_allow_html=True)
                _render_bt(usa_bt, "usa_bt")

        # ── TARAMA SONUÇLARI ───────────────────────────────────────────────
        if bist_scan or usa_scan:
            st.markdown("---")
            st.markdown("### 🔍 Tarama Sonuçları")
            if bist_scan:
                st.markdown(
                    "<h4 style='color:#f59e0b;border-left:4px solid #f59e0b;padding-left:0.6rem;'>🇹🇷 BIST Tarama Sonuçları</h4>",
                    unsafe_allow_html=True)
                _render_scan(bist_scan, "bist_scan")
            if usa_scan:
                st.markdown(
                    "<h4 style='color:#3b82f6;border-left:4px solid #3b82f6;padding-left:0.6rem;'>🇺🇸 USA Tarama Sonuçları</h4>",
                    unsafe_allow_html=True)
                _render_scan(usa_scan, "usa_scan")

with tab_guide:
    st.markdown("## Doğru Kullanım Şekli")
    st.markdown("---")

    st.markdown("### 📋 En İyi Kullanıcı Akışı")

    st.markdown("#### 1. Genel Fırsat Tarama")
    st.markdown(
        """
- **Piyasa:** BIST
- **Evren:** BIST100
- **Tarama Modu:** Standart Tarama
- **Temel Kalite Seviyesi:** Temel
- **Sıralama:** Combined Score
        """
    )

    st.markdown("#### 2. Akıllı Para Girişi Bulma")
    st.markdown(
        """
- **Piyasa:** BIST
- **Evren:** BISTTUM
- **Tarama Modu:** Sadece Akıllı Para Girenler
- **Temel Kalite Seviyesi:** Temel
- **Sıralama:** Combined Score
        """
    )

    st.markdown("#### 3. Erken Accumulation Yakalama")
    st.markdown(
        """
- **Piyasa:** BIST
- **Evren:** BIST100 DIŞI
- **Tarama Modu:** Erken Accumulation Yakalama
- **Temel Kalite Seviyesi:** Kapalı veya Temel
- **Sıralama:** Technical Score veya Combined Score
        """
    )

    st.markdown("---")
    st.markdown("### 🔢 En Doğru Sıralama Mantığı")

    st.markdown("#### Standart Tarama")
    st.markdown("- Varsayılan sıralama: **Combined Score**")

    st.markdown("#### Akıllı Para Girenler")
    st.markdown("- Önerilen sıralama: **Combined Score** veya **Technical Score**")

    st.markdown("#### Erken Accumulation")
    st.markdown("- Önerilen sıralama: **Technical Score** veya **Combined Score**")

    st.info("İlk kullanımda **Combined Score** sıralaması önerilir.")

    st.markdown("---")
    st.markdown("### 🧠 Temel Mantık")
    st.markdown(
        """
- **Tarama Modu** → piyasadaki davranışı seçer
- **Temel Kalite Seviyesi** → şirket kalitesini filtreler
- **Sıralama Türü** → sonuçları hangi kritere göre dizdiğini belirler
        """
    )

    st.markdown("---")
    st.markdown("### 💡 Kullanım Tavsiyeleri")
    st.markdown(
        """
- Çok sıkı filtreler az sonuç getirir
- Erken fırsatlar için kalite filtresi gevşetilebilir
- En iyi sonuçlar **Combined Score** ile alınır
- **Backtest** ile strateji doğrulaması yapılmalıdır
        """
    )
